"""
CTDoseExcel.py - Conversor de dados de relatórios CT para Excel

Este script lê os arquivos JSON gerados pelo CTDoseExtractor e cria
uma planilha Excel com as informações organizadas em colunas.
"""

import json
import os
import glob
import argparse
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def calculate_age(birth_date_str, exam_date_str):
    """Calcula a idade com base na data de nascimento e data do exame"""
    if not birth_date_str or not exam_date_str:
        return '-'

    # Extrai o ano da data de nascimento
    birth_year_match = re.search(r'(\d{4})', birth_date_str)
    if not birth_year_match:
        # Tenta outro formato: "Jul 1, 1997"
        month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        for month in month_names:
            if month in birth_date_str:
                parts = birth_date_str.replace(",", "").split()
                if len(parts) >= 3 and parts[2].isdigit() and len(parts[2]) == 4:
                    birth_year = int(parts[2])
                    break
        else:
            return '-'  # Não conseguiu extrair o ano
    else:
        birth_year = int(birth_year_match.group(1))

    # Extrai o ano da data do exame
    exam_year_match = re.search(r'(\d{4})', exam_date_str)
    if not exam_year_match:
        return '-'  # Não conseguiu extrair o ano

    exam_year = int(exam_year_match.group(1))

    # Calcula a idade simples (apenas com base nos anos)
    # Este é um cálculo aproximado sem considerar mês/dia
    age = exam_year - birth_year

    return str(age)


def extract_scan_info(acquisition):
    """Extrai informações de aquisição para cada linha, exatamente como estão no JSON"""
    scan_info = {}

    # Protocolo (será usado como Pesquisa de interesse)
    scan_info['protocol'] = acquisition.get('protocol', '-')

    # Descrição da série (APENAS comment, sem fallback)
    # Tratamento especial para comment (garantir que null se torne '-')
    comment = acquisition.get('comment')
    # Múltiplas verificações para garantir que qualquer valor "vazio" se torne '-'
    if comment is None or comment == '' or (isinstance(comment, str) and comment.strip() == '') or comment == 'null':
        scan_info['description'] = '-'
    else:
        scan_info['description'] = comment

    # Scan mode (tipo de aquisição)
    scan_info['scan_mode'] = acquisition.get('acquisition_type', '-')

    # Phantom type
    ct_dose = acquisition.get('ct_dose', {}) or {}  # Usa {} se for None
    scan_info['phantom_type'] = ct_dose.get('phantom_type', '-')

    # CTDI vol - valor exato como está no JSON
    ctdivol = ct_dose.get('mean_ctdivol')
    scan_info['ctdivol'] = ctdivol if ctdivol is not None else '-'

    # DLP - valor exato como está no JSON
    dlp = ct_dose.get('dlp')
    scan_info['dlp'] = dlp if dlp is not None else '-'

    # SSDE - valor exato como está no JSON, com tratamento especial para None
    ssde = ct_dose.get('size_specific_dose')
    scan_info['ssde'] = ssde if ssde is not None else '-'

    # Dados da fonte de raios X
    xray_params = acquisition.get('xray_source_params', {}) or {}  # Usa {} se for None

    # Tube current - valor exato como está no JSON
    tube_current = xray_params.get('tube_current')
    scan_info['tube_current'] = tube_current if tube_current is not None else '-'

    # kV - valor exato como está no JSON
    kv = xray_params.get('kvp')
    scan_info['kv'] = kv if kv is not None else '-'

    # Avg scan size - não está disponível geralmente
    scan_info['avg_scan_size'] = '-'

    return scan_info


def json_to_excel(json_folder="ct_reports_json", output_file="ct_dose_report.xlsx"):
    """Converte os dados JSON para Excel"""
    # Encontra o arquivo JSON coletivo
    json_all_path = os.path.join(json_folder, "ct_reports_all.json")

    if not os.path.exists(json_all_path):
        # Tenta encontrar qualquer arquivo JSON na pasta
        json_files = glob.glob(os.path.join(json_folder, "*.json"))
        if not json_files:
            print(f"❌ Erro: Nenhum arquivo JSON encontrado em '{json_folder}'")
            return False

        # Usa o primeiro arquivo JSON encontrado
        json_all_path = json_files[0]
        print(f"ℹ️ Arquivo ct_reports_all.json não encontrado, usando '{os.path.basename(json_all_path)}'")

    # Lê o arquivo JSON
    try:
        with open(json_all_path, 'r', encoding='utf-8') as f:
            reports = json.load(f)

        if not isinstance(reports, list):
            reports = [reports]  # Converte para lista se for apenas um objeto

        print(f"✓ Leitura de {len(reports)} relatórios do arquivo '{os.path.basename(json_all_path)}'")
    except Exception as e:
        print(f"❌ Erro ao ler JSON: {str(e)}")
        return False

    # Cria uma nova planilha Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatórios CT"

    # Define os cabeçalhos
    headers = [
        "ID do paciente", "Sexo", "Data de nascimento", "Idade", "Pesquisa de interesse",
        "Data do exame", "Descrição da série", "Scan mode", "mAs",
        "kV", "CTDIvol", "DLP", "DLP total", "Phantom type", "SSDE", "Avg scan size"
    ]

    # Formata os cabeçalhos
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Adiciona cabeçalhos na primeira linha
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Define larguras das colunas
    ws.column_dimensions['A'].width = 15  # ID do paciente
    ws.column_dimensions['B'].width = 10  # Sexo
    ws.column_dimensions['C'].width = 18  # Data de nascimento
    ws.column_dimensions['D'].width = 10  # Idade
    ws.column_dimensions['E'].width = 20  # Pesquisa de interesse
    ws.column_dimensions['F'].width = 18  # Data do exame
    ws.column_dimensions['G'].width = 20  # Descrição da série
    ws.column_dimensions['H'].width = 15  # Scan mode
    ws.column_dimensions['I'].width = 10  # mAs
    ws.column_dimensions['J'].width = 10  # kV
    ws.column_dimensions['K'].width = 10  # CTDIvol
    ws.column_dimensions['L'].width = 10  # DLP
    ws.column_dimensions['M'].width = 10  # DLP total
    ws.column_dimensions['N'].width = 15  # Phantom type
    ws.column_dimensions['O'].width = 10  # SSDE
    ws.column_dimensions['P'].width = 15  # Avg scan size

    # Linha atual para inserção
    row_idx = 2

    # Processa cada relatório
    for report in reports:
        essential = report.get('essential', {})

        # Obtém informações básicas do paciente/estudo
        patient_id = essential.get('patient_id', '')
        sex = essential.get('sex', '')
        birth_date = essential.get('birth_date', '')  # Mantemos a data de nascimento, sem calcular idade
        study_date = essential.get('study_date', '')

        # DLP total - direto do JSON, sem extração
        irradiation = report.get('irradiation', {})
        total_dlp = irradiation.get('total_dlp', '')

        # Processa cada aquisição/série como uma linha
        acquisitions = report.get('acquisitions', [])
        if acquisitions:
            for acquisition in acquisitions:
                # Obtém informações desta aquisição específica
                scan_info = extract_scan_info(acquisition)

                # Insere valores na planilha com tratamento explícito para None/null
                ws.cell(row=row_idx, column=1, value=patient_id if patient_id is not None else '-')
                ws.cell(row=row_idx, column=2, value=sex if sex is not None else '-')
                ws.cell(row=row_idx, column=3, value=birth_date if birth_date is not None else '-')

                # Calcula a idade com base na data de nascimento e data do exame
                age = calculate_age(birth_date, study_date)
                ws.cell(row=row_idx, column=4, value=age)

                ws.cell(row=row_idx, column=5, value=scan_info['protocol'])
                ws.cell(row=row_idx, column=6, value=study_date if study_date is not None else '-')
                # Descrição da série - tratamento especial para garantir '-' em caso de null
                description_value = scan_info['description']
                # Verificação extra rigorosa para garantir que não seja null, string vazia, espaços, etc.
                is_empty = (description_value is None or description_value == '' or
                            description_value.strip() == '' or description_value == 'null')
                ws.cell(row=row_idx, column=7, value='-' if is_empty else description_value)
                ws.cell(row=row_idx, column=8, value=scan_info['scan_mode'])
                ws.cell(row=row_idx, column=9, value=scan_info['tube_current'])
                ws.cell(row=row_idx, column=10, value=scan_info['kv'])
                ws.cell(row=row_idx, column=11, value=scan_info['ctdivol'])
                ws.cell(row=row_idx, column=12, value=scan_info['dlp'])
                ws.cell(row=row_idx, column=13, value=total_dlp if total_dlp is not None else '-')
                ws.cell(row=row_idx, column=14, value=scan_info['phantom_type'])
                ws.cell(row=row_idx, column=15, value=scan_info['ssde'])
                ws.cell(row=row_idx, column=16, value=scan_info['avg_scan_size'])

                # Aplica borda a todas as células
                for col_idx in range(1, 17):
                    ws.cell(row=row_idx, column=col_idx).border = border

                row_idx += 1
        else:
            # Se não houver aquisições, adiciona pelo menos uma linha com dados básicos
            ws.cell(row=row_idx, column=1, value=patient_id if patient_id is not None else '-')
            ws.cell(row=row_idx, column=2, value=sex if sex is not None else '-')
            ws.cell(row=row_idx, column=3, value=birth_date if birth_date is not None else '-')

            # Calcula a idade para esta linha também
            age = calculate_age(birth_date, study_date)
            ws.cell(row=row_idx, column=4, value=age)

            ws.cell(row=row_idx, column=5, value='-')
            ws.cell(row=row_idx, column=6, value=study_date if study_date is not None else '-')
            ws.cell(row=row_idx, column=7, value='-')
            ws.cell(row=row_idx, column=8, value='-')
            ws.cell(row=row_idx, column=9, value='-')
            ws.cell(row=row_idx, column=10, value='-')
            ws.cell(row=row_idx, column=11, value='-')
            ws.cell(row=row_idx, column=12, value='-')
            ws.cell(row=row_idx, column=13, value=total_dlp if total_dlp is not None else '-')
            ws.cell(row=row_idx, column=14, value='-')
            ws.cell(row=row_idx, column=15, value='-')
            ws.cell(row=row_idx, column=16, value='-')

            # Aplica borda a todas as células
            for col_idx in range(1, 17):
                ws.cell(row=row_idx, column=col_idx).border = border

            row_idx += 1

    # Salva a planilha
    try:
        wb.save(output_file)
        print(f"✅ Planilha Excel salva com sucesso: '{output_file}'")
        print(f"   Total de linhas geradas: {row_idx - 2}")
        return True
    except Exception as e:
        print(f"❌ Erro ao salvar planilha Excel: {str(e)}")
        return False


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Conversor de relatórios JSON de CT para Excel')
    parser.add_argument('--input-folder', type=str, default='ct_reports_json',
                        help='Pasta contendo os arquivos JSON (padrão: ct_reports_json)')
    parser.add_argument('--output', type=str, default='ct_dose_report.xlsx',
                        help='Nome do arquivo Excel de saída (padrão: ct_dose_report.xlsx)')

    args = parser.parse_args()

    print(f"\n{'=' * 80}")
    print(f"CTDoseExcel - Conversor de Relatórios JSON para Excel")
    print(f"{'=' * 80}")
    print(f"Pasta de entrada (JSONs): {args.input_folder}")
    print(f"Arquivo de saída (Excel): {args.output}")
    print(f"{'=' * 80}\n")

    json_to_excel(args.input_folder, args.output)